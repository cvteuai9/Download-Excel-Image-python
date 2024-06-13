import openpyxl
import requests
from pathlib import Path

def download_images_from_excel(excel_path, column_letter, output_folder, filename_column_letter):
    # 加載 Excel 文件
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active  # 獲取第一个工作表

    # 創建保存圖片的文件夾
    Path(output_folder).mkdir(parents=True, exist_ok=True)

    # 讀取 Excel 表格中的每一行，獲取指定列的 URL
    serial_number = 1
    # min_row = 從sheet的第幾列開始下載
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        cell = row[ord(column_letter.upper()) - ord('A')]
        url = cell.value
        filename_cell = row[ord(filename_column_letter.upper()) - ord('A')]

        if url and isinstance(url, str) and url.startswith('http'):
            try:
                # 生成图片文件名（流水号）
                filename = f"image_{serial_number:04d}.jpg"  # 例如 image_0001.jpg
                filename_cell.value = filename

                # 下載圖片
                response = requests.get(url)
                response.raise_for_status()  # 如果請求出錯，拋出異常

                # 生成圖片文件名（流水號）
                output_path = Path(output_folder) / filename

                # 保存圖片
                with open(output_path, 'wb') as file:
                    file.write(response.content)
                
                print(f"下載成功: {filename}")
                serial_number += 1
            except requests.RequestException as e:
                print(f"下載失敗: {url} - {e}")

    # 保存修改后的 Excel 文件
    workbook.save(excel_path)
    print(f"Excel 文件已更新并保存到 {excel_path}")

if __name__ == "__main__":
    excel_path = 'images.xlsx'  # Excel 文件路徑
    column_letter = 'A'  # 儲存 URL 的行，如 'A'
    output_folder = 'images'  # 圖片保存的文件夹
    filename_column_letter = 'B'  # 存储文件名的列，如 'B'
    
    download_images_from_excel(excel_path, column_letter, output_folder, filename_column_letter)

import openpyxl
from pathlib import Path

def update_excel_with_filenames(excel_path, url_column_letter, filename_column_letter):
    # 加载 Excel 文件
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active  # 获取第一个工作表

    # 初始化流水号
    serial_number = 1

    # 读取 Excel 表格中的每一行，处理指定列的 URL，并填入文件名
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        url_cell = row[ord(url_column_letter.upper()) - ord('A')]
        filename_cell = row[ord(filename_column_letter.upper()) - ord('A')]

        if url_cell.value and isinstance(url_cell.value, str) and url_cell.value.startswith('http'):
            # 生成图片文件名（流水号）
            filename = f"image_{serial_number:04d}.jpg"  # 例如 image_0001.jpg
            filename_cell.value = filename
            serial_number += 1

    # 保存修改后的 Excel 文件
    workbook.save(excel_path)
    print(f"Excel 文件已更新并保存到 {excel_path}")

if __name__ == "__main__":
    excel_path = 'images.xlsx'  # Excel 文件路径
    url_column_letter = 'A'  # 存储 URL 的列，如 'A'
    filename_column_letter = 'B'  # 存储文件名的列，如 'B'

    update_excel_with_filenames(excel_path, url_column_letter, filename_column_letter)
